from tkinter import filedialog
from langdetect import detect, DetectorFactory
from docx import Document
import pandas as pd
import openpyxl
import string
from titlecase import titlecase
from datetime import datetime, timedelta
from docx.opc.constants import RELATIONSHIP_TYPE as RT
import re
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
import validators


template="C:\\Users\\desjardav\\Downloads\\VLB_Content_matrix_SFXXX.xlsx"
DetectorFactory.seed = 0

alphabet=list(string.ascii_lowercase)

months={
"janvier":"01",
"février":"02",
"mars":"03",
"avril":"04",
"mai":"05",
"juin":"06",
"juillet":"07",
"août":"08",
"septembre":"09",
"octobre":"10",
"décembre":"11",
}

def remove_hyperlink_tags(xml):
    #text = xml.decode('utf-8')
    text = xml.replace("</w:hyperlink>","")
    text = re.sub('<w:hyperlink[^>]*>', "", text)
    return text.encode('utf-8')

datatree={
    0:[[0,0]],
    1:[[0,0]],
    2:[[0,0]],
    3:[[]],
    4:[[0,1],[0,2],[0,3],[0,4]],
    5:[[0,1]],
    6:[[0,0],[0,1],[0,2]],
    7:[[0,0],[0,1],[0,2]],
    8:[[0,1],[0,2]],
    9:[[0,1]],
    10:[[0,0],[0,1]],
    11:[[0,0],[0,1]],
    12:[[]],
    13:[[0,1],[0,2],[0,3],[0,4]],
    }

def clean_date(leafytree):
    entry=leafytree[0][0]
    entry=entry.lower()
    for key in months.keys():
        if key in entry:
            entry=entry.replace(key,"/"+months[key]+"/")
            break
    
    for letter in list(entry):
        if letter not in ["0","1","2","3","4","5","6","7","8","9","/"]:
            entry=entry.replace(letter,"")
    entry = datetime.strptime(entry,"%d/%m/%Y") - timedelta(days=8)
    leafytree[0][0]=entry.date()
    return leafytree

def check_punctuation(leafytree):
    targets=[[1,0],[2,0]]
    
    for target in targets:
        print(leafytree)
        string=leafytree[target[0]][target[1]]
        if string[len(string)-1] not in ["?","!","."]:
            string=string+"."
        else:
            string=string
        leafytree[target[0]][target[1]]=string
    return leafytree

def capitalize(leafytree):
    ## [datatree key, position in datatree list]
    targets=[[1,0],[2,0],[4,0],[5,0],[6,0],[7,0],[8,0],[9,0],[10,0],[11,0],[13,0]]
    for target in targets:
        entry=titlecase(leafytree[target[0]][target[1]])
        leafytree[target[0]][target[1]]=entry
    return leafytree


def pulltree(target):
    wordDoc=Document(target)

    counter=0
    leafytree={}
    for table in wordDoc.tables:
        minilist=[]
        for coordinates in datatree[counter]:
            if coordinates==[]:
                continue
            content=table.cell(coordinates[0],coordinates[1]).text
            print(content)
            allxml=table.cell(coordinates[0],coordinates[1])._element.xml
            allxml=remove_hyperlink_tags(allxml)
            soup=BeautifulSoup(allxml)
            dirtyleaf=soup.findAll(text=True)
            #dirtyleaf=ET.tostring(ET.fromstring(allxml), encoding='utf-8', method='text')
            removelist=["",None,"\n"]#"b'\\n","\\n'"]
            
            for i in removelist:
                for x in range(dirtyleaf.count(i)):
                    dirtyleaf.remove(i)
            print(dirtyleaf,counter)
            
            exception_handler=[counter,coordinates]
            exceptions=[[0,[0,0]],[6,[0,1]],[7,[0,1]]]

            if len(dirtyleaf)>=1:
                if exception_handler in exceptions:
                    appendthis=dirtyleaf[2]
                else:
                    appendthis=dirtyleaf[1]
            else: appendthis=dirtyleaf[0]
            if validators.url(appendthis)!=True:
                if "\n" in content:
                    appendthis=content.split("\n")[1]
                    
            minilist.append(appendthis)
        leafytree[counter]=minilist
        counter+=1
    return clean_date(check_punctuation(leafytree))

targets=filedialog.askopenfilenames()
treelang={}

for target in targets:
    
    leafytree=pulltree(target)
    lang=detect(leafytree[13][1]+leafytree[4][1])
    if lang=="fr":
        treelang["fr"]=leafytree
    elif lang=="en":
        treelang["en"]=capitalize(leafytree)

xfile = openpyxl.load_workbook(template)


def WriteData():

    for key in treelang.keys():
        if key=="en":
            sheetthread="Data_EN"
        elif key=="fr":
            sheetthread="Data_FR"
        sheet=xfile[sheetthread]
        sheet.delete_cols(1,10)
        sheet.delete_rows(1,1000)
        counterrow=1
        for i in treelang[key]:
            counter=0
            if treelang[key][i]==[] or treelang[key][i]==['']:
                continue
            for x in treelang[key][i]:
                if x=="" or x==None:
                    continue
                col=alphabet[counter]
                sheet[col+str(counterrow)]=x
                counter+=1
            counterrow+=1
    xfile.save(template)

WriteData()








